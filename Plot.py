import matplotlib
import matplotlib.font_manager
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from PyQt5 import QtCore 
from PyQt5.QtGui import QImage, QPixmap
import cv2 as cv

class Plot:
    def __init__(self, ui):
        self.ui = ui
        self.qpixmap_fix_width = 500
        self.qpixmap_fix_height = 380
        self.wb = openpyxl.load_workbook('Result.xlsx')
        self.ws = self.wb.worksheets[0]
        matplotlib.rcParams['font.family'] = ['Microsoft JhengHei']    # 讓字形可以用中文字

    def judge_fiveDays(self):
        if(self.ws['A6'].value):                    # 有5天的資料了
            return True
        else:
            return False

    # 從Excel讀資料
    def load_currentFiveDaysData(self):

        x = [1, 2, 3, 4, 5]                     # x軸資料
        dates_labels = [self.ws['A2'].value, self.ws['A3'].value, self.ws['A4'].value, self.ws['A5'].value, self.ws['A6'].value]                    # x軸標籤名稱   

        # y軸資料
        total_surgeyTime = [self.ws['B2'].value, self.ws['B3'].value, self.ws['B4'].value, self.ws['B5'].value, self.ws['B6'].value]
        total_interruptTime = [self.ws['C2'].value, self.ws['C3'].value, self.ws['C4'].value, self.ws['C5'].value, self.ws['C6'].value]
        total_interruptCounts = [self.ws['D2'].value, self.ws['D3'].value, self.ws['D4'].value, self.ws['D5'].value, self.ws['D6'].value]
        ratio_of_intereuptTime = [self.ws['E2'].value, self.ws['E3'].value, self.ws['E4'].value, self.ws['E5'].value, self.ws['E6'].value]
        averageTime_of_oneInterrupt = []

        self.plot_graph1(x, total_surgeyTime, dates_labels)
        self.plot_graph2(x, total_interruptTime, dates_labels)
        self.plot_graph3(x, total_interruptCounts, dates_labels)
        self.plot_graph4(x, ratio_of_intereuptTime, dates_labels)



    # graph1: 總手術時間
    def plot_graph1(self, x, total_surgeyTime, dates_labels):

        # 中文版
        fig, ax = plt.subplots()
        bar_container = ax.bar(x, total_surgeyTime, color='cornflowerblue', tick_label=dates_labels, width=0.4)
        ax.set_xlabel('手術日期')
        ax.set_ylabel('時間 (分鐘)')
        ax.set_title('總手術時間')
        ax.bar_label(bar_container)
        
        plt.savefig('./result_image/Img1.png')
        plt.close()

        # 將圖片顯示在label上
        img1 = cv.imread('./result_image/Img1.png')
        height, width, channel = img1.shape

        bytes_perline = 3 * width
        qimage = QImage(img1, width, height, bytes_perline, QImage.Format_RGB888).rgbSwapped()
        self.ui.plot_label1.setPixmap(QPixmap.fromImage(qimage))


        # # 英文版
        # plt.xlabel('Sugery\'s Date')
        # plt.ylabel('Time (min)')
        # plt.title('Total Interrupt Time of a Date')

    # graph2: 總中段時間
    def plot_graph2(self, x, total_interruptTime, dates_labels):

        # 中文版
        fig, ax = plt.subplots()
        bar_container = ax.bar(x, total_interruptTime, color='cornflowerblue', tick_label=dates_labels, width=0.4)
        ax.set_xlabel('手術日期')
        ax.set_ylabel('時間 (分鐘)')
        ax.set_title('總中斷時間')
        ax.bar_label(bar_container)
        
        plt.savefig('./result_image/Img2.png')
        plt.close()

        # 將圖片顯示在label上
        img1 = cv.imread('./result_image/Img2.png')
        height, width, channel = img1.shape

        bytes_perline = 3 * width
        qimage = QImage(img1, width, height, bytes_perline, QImage.Format_RGB888).rgbSwapped()
        self.ui.plot_label2.setPixmap(QPixmap.fromImage(qimage))


    # graph3: 總中斷次數
    def plot_graph3(self, x, total_interruptCounts, dates_labels):

        # 中文版
        fig, ax = plt.subplots()
        bar_container = ax.bar(x, total_interruptCounts, color='cornflowerblue', tick_label=dates_labels, width=0.4)
        ax.set_xlabel('手術日期')
        ax.set_ylabel('次數')
        ax.set_title('總中斷次數')
        ax.bar_label(bar_container)
        
        plt.savefig('./result_image/Img3.png')
        plt.close()

        # 將圖片顯示在label上
        img1 = cv.imread('./result_image/Img3.png')
        height, width, channel = img1.shape

        bytes_perline = 3 * width
        qimage = QImage(img1, width, height, bytes_perline, QImage.Format_RGB888).rgbSwapped()
        self.ui.plot_label3.setPixmap(QPixmap.fromImage(qimage))

        
        
    # graph4: 中斷時間/總手術時間的比例
    def plot_graph4(self, x, ratio_of_intereuptTime, dates_labels):

        # 中文版
        fig, ax = plt.subplots()
        bar_container = ax.bar(x, ratio_of_intereuptTime, color='cornflowerblue', tick_label=dates_labels, width=0.4)
        ax.set_xlabel('手術日期')
        ax.set_ylabel('百分比 (%)')
        ax.set_title('中斷時間/總手術時間的比例 (%)')
        ax.bar_label(bar_container)
        
        plt.savefig('./result_image/Img4.png')
        plt.close()

        # 將圖片顯示在label上
        img1 = cv.imread('./result_image/Img4.png')
        height, width, channel = img1.shape

        bytes_perline = 3 * width
        qimage = QImage(img1, width, height, bytes_perline, QImage.Format_RGB888).rgbSwapped()
        self.ui.plot_label4.setPixmap(QPixmap.fromImage(qimage))
        

    # # graph5: 平均一次中斷花的時間
    # def plot_graph5(x, averageTime_of_oneInterrupt, dates_labels):
        



    