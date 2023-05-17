import numpy as np
import openpyxl
import os
from Utils import compute
from Utils import judge

class Directory:
    def __init__(self, dirpath, dirname):
        # 一開始就得初始化好的
        self.dirpath = dirpath
        self.dirname = dirname
        self.video_count = 0
        self.oneday_total_time = 0
        self.video_file_list = []

        # 經過judge後才能找到的
        self.oneday_interrupt_count = 0
        self.oneday_interrupt_time = 0
        self.oneday_performance = '-'
        self.oneday_efficiency = '-'
        self.oneday_ratio = 0
        self.oneday_unit_interrupt_counts = 0

        self.wb = openpyxl.load_workbook('Result.xlsx')

    # 計算一天的手術資訊  ex.總手術時間、總中斷時間、總中斷次數.....
    def calculate_onedayInfo(self):

        for video in self.video_file_list:
            for i in range(video.total_revised_interrupt_count):
                interrupt = video.revised_interrupt_list[i]

                self.oneday_interrupt_time += interrupt["end_time"] - interrupt["start_time"]
                self.oneday_interrupt_count += 1

            # self.oneday_total_time = 

        self.oneday_ratio = round((100 * self.oneday_interrupt_time/self.oneday_total_time), 1)
        self.oneday_unit_interrupt_counts = round(self.oneday_interrupt_count/(self.oneday_total_time/60/9), 2)
        self.oneday_performance = judge.performance(oneday_dir=self)
        self.oneday_efficiency = judge.performance_eff(oneday_dir=self)


    def write_result_to_excel(self): ## write result to excel ##
        self.ws = self.wb.create_sheet(self.dirname)
        self.ws['A1'] = self.dirname
        self.ws['B1'] = 'Start Time'
        self.ws['C1'] = 'End Time'

        ##### 一. 印出同一天手術之不同影片的所有interrupt #####
        for video in self.video_file_list:
            data = ['', '', '']
            self.ws.append(data)

            data = [video.filename, '', '']
            self.ws.append(data)

            if video.total_revised_interrupt_count == 0:
                data = ['No interrupt', '', '']
                self.ws.append(data)

            for i in range(video.total_revised_interrupt_count):
                interrupt_name = 'Interrupt#' + str(i+1)
                interrupt = video.revised_interrupt_list[i]
                
                ##### interrupt start #####
                start_normal_time = compute.get_normal_time_info(time_in_seconds=interrupt["start_time"])
                start_time_name = compute.get_excel_str(normal_time=start_normal_time)

                ##### interrupt end #####
                end_normal_time = compute.get_normal_time_info(time_in_seconds=interrupt["end_time"])
                end_time_name = compute.get_excel_str(normal_time=end_normal_time)

                # 加入row
                data = [interrupt_name, start_time_name, end_time_name]
                self.ws.append(data)

        for i in range(2):
            self.ws.append(['', '', ''])

        ##### 二、把手術的整體數據印在該天的excel sheet中 #####
        ### 1. 印出總手術時間 ###
        total_normal_time = compute.get_normal_time_info(time_in_seconds=self.oneday_total_time)
        total_time_name = compute.get_excel_str(normal_time=total_normal_time)
        data = ['手術總執行時間', total_time_name, '']
        self.ws.append(data)

        ### 2. 印出手術總中斷時間 ###
        total_interrupt_normal_time = compute.get_normal_time_info(time_in_seconds=self.oneday_interrupt_time)
        total_interrupt_time_name = compute.get_excel_str(normal_time=total_interrupt_normal_time)
        data = ['手術總中斷時間', total_interrupt_time_name, '']
        self.ws.append(data)

        ### 3. 印出手術總中斷次數 ###
        total_interrupt_number_name = str(self.oneday_interrupt_count)
        data = ['手術總中斷次數', total_interrupt_number_name, '']
        self.ws.append(data)

        ### 4.印出中斷時間佔整個手術時間的比例 ###
        ratio_name = str(self.oneday_ratio)
        data = ['中斷時間佔整個手術時間的比例', ratio_name, '']
        self.ws.append(data)

        ### 5. 印出平均每九分鐘的中斷次數 ###
        oneday_unit_interrupt_counts = str(self.oneday_unit_interrupt_counts)
        data = ['平均每九分鐘的中斷次數', oneday_unit_interrupt_counts, '']
        self.ws.append(data)

        ### 6. 印出performance ###
        data = ['手術中斷次數評分', self.oneday_performance, '']
        self.ws.append(data)

        ### 7. 印出efficiency ###
        data = ['手術效率評分', self.oneday_efficiency, '']
        self.ws.append(data)


        ##### 三. 將這天的手術資訊印到Excel的總表 #####
        target_sheetName = '目前測過的手術日期總資訊'
        exist_flag = False
        for i in range(len(self.wb.sheetnames)):
            if(self.wb.sheetnames[i] == target_sheetName):
                exist_flag = True
                break
        
        if(exist_flag == False):        # sheet不存在就建一個
            self.wb.create_sheet("目前測過的手術日期總資訊", 0)
            ws1 = self.wb.worksheets[0]
            data = ['手術日期', '總手術時間 (分)', '總中斷時間 (分)', '總中斷次數', '中斷時間/總手術時間的比例 (%)', '手術表現評分', '手術效率評分']
            ws1.append(data)
        else:
            ws1 = self.wb.worksheets[0]

        total_surgeryTime = round(self.oneday_total_time/60, 1)         #分鐘數
        total_interruptTime = round(self.oneday_interrupt_time/60, 1)   #分鐘數
        data = [self.dirname, total_surgeryTime, total_interruptTime, self.oneday_interrupt_count, self.oneday_ratio, self.oneday_performance, self.oneday_efficiency]
        
        ws1.insert_rows(2)
        ws1._current_row = 1
        ws1.append(data)

        ##### 四. 存檔 #####
        self.wb.save(filename='Result.xlsx')



    def delete_excel_row(self, choose_video_file, remove_interrupt_index):
        self.ws = self.wb[self.dirname]
        remove_ws = self.wb['remove']
        video_file_name = self.ws['A']

        # 直接刪除excel檔案裡的中斷
        row_num = 0
        for name in video_file_name:
            row_num += 1
            if name.value == choose_video_file.filename:
                remove_row = row_num + remove_interrupt_index + 1
                data = [choose_video_file.filename, '', '']
                remove_ws.append(data)
                row_with_values = [cell.value for cell in self.ws[str(remove_row)]]
                remove_ws.append(row_with_values)
                self.ws.delete_rows(remove_row) #excel第一列為1(非index)
                break

        # 對變數進行更改
        self.oneday_interrupt_count -= 1
        self.oneday_interrupt_time -= (choose_video_file.revised_interrupt_list[remove_interrupt_index]["end_time"] - choose_video_file.revised_interrupt_list[remove_interrupt_index]["start_time"])
        self.oneday_ratio = round((100 * self.oneday_interrupt_time/self.oneday_total_time), 1)
        self.oneday_unit_interrupt_counts = round(self.oneday_interrupt_count/(self.oneday_total_time/60/9), 2)
        self.oneday_performance = judge.performance(oneday_dir=self)
        self.oneday_efficiency = judge.performance_eff(oneday_dir=self)

        total_normal_time = compute.get_normal_time_info(time_in_seconds=self.oneday_interrupt_time)
        total_time_name = compute.get_excel_str(normal_time=total_normal_time)
        ratio_name = str(self.oneday_ratio)
        oneday_unit_interrupt_counts = str(self.oneday_unit_interrupt_counts)

        self.ws['B'+str(self.ws.max_row - 5)].value = total_time_name
        self.ws['B'+str(self.ws.max_row - 4)].value = self.oneday_interrupt_count
        self.ws['B'+str(self.ws.max_row - 3)].value = ratio_name
        self.ws['B'+str(self.ws.max_row - 2)].value = oneday_unit_interrupt_counts
        self.ws['B'+str(self.ws.max_row - 1)].value = self.oneday_performance
        self.ws['B'+str(self.ws.max_row)].value = self.oneday_efficiency


        self.wb.save(filename='Result.xlsx')
        