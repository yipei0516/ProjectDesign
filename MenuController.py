from PyQt5 import QtWidgets, QtGui, QtCore
from ResultController import Result_controller
from PlotController import Plot_controller
from Ui_Menu import Ui_Form
from Directory import Directory
from File import Video_File
from Utils import judge, compute, opencv_engine, image
from Plot import Plot
import cv2 as cv 
import os
from playsound import playsound
import openpyxl


class Menu_controller(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.setup_control()
        self.setWindowFlags(QtCore.Qt.WindowMinimizeButtonHint | QtCore.Qt.WindowCloseButtonHint)

        qimage = image.show_image_on_label("./image/s2.png")
        self.ui.label_background.setPixmap(qimage)
        self.ui.label_background.lower()
        op = QtWidgets.QGraphicsOpacityEffect()
        op.setOpacity(0.65)
        self.ui.label_background.setGraphicsEffect(op)

        qimage = image.show_image_on_label("./image/shine.png")
        self.ui.label_title_image.setPixmap(qimage)
        self.ui.label_title_image_2.setPixmap(qimage)

        self.ui.list_widget_done_folder.hide()
        self.ui.button_back_menu.hide()

        self.ui.button_start_judge.setDisabled(True)

        self.wb = openpyxl.load_workbook('Result.xlsx')


    def setup_control(self):
        self.ui.button_choose_video.clicked.connect(self.clicked_choose_video)
        self.ui.button_start_judge.clicked.connect(self.clicked_start_judge)
        self.ui.button_plot_result.clicked.connect(self.clicked_button_plot_result)
        self.ui.button_manual.clicked.connect(self.show_manual)
        self.ui.button_manual_2.clicked.connect(self.show_manual)
        self.ui.button_check_folder.clicked.connect(self.check_done_folder)
        self.ui.button_back_menu.clicked.connect(self.back_menu)
        


    def clicked_choose_video(self):
        ##### Step1. 選取資料夾 #####
        dir_path = QtWidgets.QFileDialog.getExistingDirectory()
        dir_name = os.path.basename(dir_path)
        self.oneday_dir = Directory(dirpath=dir_path, dirname=dir_name)


        ##### Step3. 確認是否直接關掉選取畫面 #####
        # 直接關掉選擇資料夾畫面 #
        if dir_path == '':
            pass
        # 已選取某個資料夾!! #
        else:
            ##### Step3. 先做Directory的部分-> 找出全部的video(小檔案) #####
            all_file_name = os.listdir(dir_path)
            oneday_video_name = []
            for file in all_file_name:
                if os.path.splitext(file)[1] == ".mp4": # 若是影片檔才能當作一個file
                    # 1. 先新增成多個video_file檔
                    filepath = dir_path + "/" + file
                    filename = os.path.splitext(file)[0]
                    video_file = Video_File(filepath=filepath, filename=filename)
                    self.oneday_dir.video_file_list.append(video_file)
                    self.oneday_dir.video_count += 1
                    self.oneday_dir.oneday_total_time += round(video_file.videoinfo['frame_count']/video_file.videoinfo['fps'], 1) # 記錄總影片時長(in seconds)

                    # 2. 再準備待會要檢查有沒有重複偵測的部分
                    oneday_video_name.append(filename)

            ##### Step4. 確認資料夾是否選對(有無包含mp4檔) #####
            # 沒有含mp4檔案 #
            if self.oneday_dir.video_count == 0:
                mbox = QtWidgets.QMessageBox()
                mbox.setIcon(QtWidgets.QMessageBox.Warning)
                mbox.setText("資料夾 {0} 未包含影片檔案，請重新選擇資料夾".format(dir_name))
                mbox.exec()
                pass
            # 有含mp4檔案 #
            else:
                ##### Step5. 確認是否偵測過此資料夾! #####
                # 工作表內已含有此天的手術偵測紀錄 #
                if dir_name in self.oneday_dir.wb.sheetnames: 
                    repeated_sheet = self.oneday_dir.wb[dir_name]
                    mbox = QtWidgets.QMessageBox()
                    mbox.setIcon(QtWidgets.QMessageBox.Warning)
                    mbox.setText("你已經測試過 {0} 的手術".format(dir_name))
                    a = mbox.addButton('重新測試', 2)
                    b = mbox.addButton('顯示上次測試結果', 2)
                    mbox.setDefaultButton(b) # 設定預設按鈕
                    ret = mbox.exec()

                    if ret == 0:
                        # 刪除前一個judge的結果
                        self.oneday_dir.wb.remove_sheet(repeated_sheet)
                        self.oneday_dir.wb.save('Result.xlsx')
                        self.ui.button_start_judge.setDisabled(False)
                    elif ret == 1:
                        # 跳出上次的結果
                        self.oneday_dir.oneday_interrupt_time = compute.normal_time_to_seconds(repeated_sheet['B'+str(repeated_sheet.max_row-5)].value)
                        self.oneday_dir.oneday_interrupt_count = int(repeated_sheet['B'+str(repeated_sheet.max_row-4)].value)
                        self.oneday_dir.oneday_ratio = repeated_sheet['B'+str(repeated_sheet.max_row-3)].value
                        self.oneday_dir.oneday_unit_interrupt_counts = repeated_sheet['B'+str(repeated_sheet.max_row-2)].value
                        self.oneday_dir.oneday_performance = repeated_sheet['B'+str(repeated_sheet.max_row-1)].value
                        self.oneday_dir.oneday_efficiency = repeated_sheet['B'+str(repeated_sheet.max_row)].value
                        
                        A_col = repeated_sheet['A']
                        start_time = repeated_sheet['B']
                        end_time = repeated_sheet['C']
                        video_index = 0
                        for i in range(2, len(A_col)): # 跳過第一行
                            if A_col[i].value == '手術總執行時間':
                                break
                            elif video_index < self.oneday_dir.video_count and A_col[i].value == oneday_video_name[video_index]:
                                video_file = self.oneday_dir.video_file_list[video_index]
                                video_index += 1
                            elif (A_col[i].value is not None) and (A_col[i].value != 'No interrupt') :
                                interrupt_info = {}
                                interrupt_info["start_time"] = compute.normal_time_to_seconds(start_time[i].value)
                                interrupt_info["end_time"] = compute.normal_time_to_seconds(end_time[i].value)
                                video_file.revised_interrupt_list.append(interrupt_info)
                                video_file.total_revised_interrupt_count += 1
                        self.ui.button_start_judge.setDisabled(True)

                        self.show_result()

                # 工作表內未含有此天的手術偵測紀錄 #
                else:
                    self.ui.button_start_judge.setDisabled(False)

                qimage = image.show_image_on_label("./image/folder.png")
                self.ui.label_folder_image.setPixmap(qimage)

                self.ui.label_video_name.setText(dir_name)


    
    ### Start Judge
    def clicked_start_judge(self):
        finish_judge = True # 是否正確結束judge
        for video in self.oneday_dir.video_file_list:
            ##### Step1. start judge #####
            finish_judge = judge.start_judge(file=video)
            if finish_judge == False: # 非正確結束
                break
            judge.revise_interrupt(file=video)

        if finish_judge == True: # 正確結束

            ##### Step2. 計算一天的手術資訊 #####
            self.oneday_dir.calculate_onedayInfo()            

            ##### Step3. write result to excel #####
            self.oneday_dir.write_result_to_excel()


            ##### Step4. 結束Judge後的提示 #####
            # 聲音 #
            playsound('./sound/done2.mp3')


            ##### Step5. 直接顯示結果 #####
            mbox = QtWidgets.QMessageBox() # 跳出警告訊息
            mbox.setIcon(QtWidgets.QMessageBox.Warning)
            mbox.setText("已結束偵測，請問")
            a = mbox.addButton('查看偵測完結果', 2)
            b = mbox.addButton('繼續偵測下一資料夾', 2)
            mbox.setDefaultButton(a) # 設定預設按鈕
            ret = mbox.exec()

            if ret == 0:
                self.show_result()
            elif ret == 1:
                self.ui.button_start_judge.setDisabled(True)
                pass


        else: # 未正確結束
            mbox = QtWidgets.QMessageBox()
            mbox.setIcon(QtWidgets.QMessageBox.Warning)
            mbox.setText("資料夾 {0} 未正確結束，請重新選擇資料夾".format(self.oneday_dir.dirname))
            self.ui.button_start_judge.setDisabled(True)
            mbox.exec()

    def show_result(self):
        self.hide()
        self.result_window = Result_controller(self.oneday_dir)
        self.result_window.show()




    def clicked_button_plot_result(self):
        self.hide()
        self.plot_window = Plot_controller()
        self.plot_window.show()


    def show_manual(self):
        print("~~")

    def check_done_folder(self):
        self.ui.list_widget_done_folder.show()
        self.ui.button_back_menu.show()
        self.ui.list_widget_done_folder.clear()
        for name in self.wb.sheetnames:
            if (name == "remove") or (name == "目前測過的手術日期總資訊"):
                pass
            else:
                self.ui.list_widget_done_folder.addItem(name)

    def back_menu(self):
        self.ui.list_widget_done_folder.hide()
        self.ui.button_back_menu.hide()
