from PyQt5 import QtWidgets, QtGui, QtCore
from UI import Ui_MainWindow
from File import Video_File
from Utils import judge, compute, opencv_engine
from VideoController import video_controller
import cv2 as cv 
import os
import re
from pydub import AudioSegment
from pydub.playback import play
import openpyxl

class MainWindow_controller(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_control()
        self.ui.button_start_judge.setDisabled(True)
        self.ui.button_check_result.setDisabled(True)

    ### 將每個button連上對應的event
    def setup_control(self):
        # TODO
        self.ui.button_choose_video.clicked.connect(self.clicked_choose_video)
        self.ui.button_start_judge.clicked.connect(self.clicked_start_judge)
        self.ui.button_check_result.clicked.connect(self.show_result)
        self.ui.list_widget_interrupt.itemDoubleClicked.connect(self.show_interrupt_clip) # 雙擊interrupt時，跳出此段畫面
        self.ui.list_widget_interrupt.itemClicked.connect(self.choose_remove_interrupt) # 單擊interrupt時，選取起來準備刪除
        self.ui.button_remove_interrupt.clicked.connect(self.remove_interrupt)

        self.ui.button_import_result.clicked.connect(self.performance)

        self.ui.comboBox_choose_video.currentIndexChanged.connect(self.change_list_widget)
    
        

    ### load video into FileDialog
    def clicked_choose_video(self):
        
        # os.chdir(os.getcwd())
        dir_path = QtWidgets.QFileDialog.getExistingDirectory()
        self.dir_name =  os.path.splitext(dir_path)[0]
        if dir_path == '':
            pass
        else:
            one_day_file_name = os.listdir(dir_path)
            self.one_day_video = []
            fn = self.dir_name + ".xlsx"
            self.wb = openpyxl.Workbook() # 創造新的workbook
            self.wb.save()
            for file in one_day_file_name:
                if os.path.splitext(file)[1] == ".mp4": # 若是影片檔才能當作一個file
                    filepath = dir_path + "/" + file
                    filename = os.path.splitext(file)[0]
                    video_file = Video_File(filepath=filepath, filename=filename)
                    self.one_day_video.append(video_file)

            self.ui.label_video_name.setText("Video Name:   " + self.dir_name)

            # if self.video_filename in self.video_file.wb.sheetnames:
            #     repeated_sheet = self.video_file.wb[self.video_filename]
            #     mbox = QtWidgets.QMessageBox(self.ui.centralwidget) # 跳出警告訊息
            #     mbox.setIcon(QtWidgets.QMessageBox.Warning)
            #     mbox.setText("你已經測試過此手術片段，是否要重新測?")
            #     # 添加三顆按鈕
            #     mbox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            #     # 設定預設按鈕
            #     mbox.setDefaultButton(QtWidgets.QMessageBox.No)
            #     ret = mbox.exec()                      # 取得點擊的按鈕數字
            #     if ret == QtWidgets.QMessageBox.Yes:
            #         # 刪除前一個judge的結果
            #         self.video_file.wb.remove_sheet(repeated_sheet)
            #         self.video_file.wb.save('Result.xlsx')
            #         print("YES")
            #     elif ret == QtWidgets.QMessageBox.No:
            #         # 跳出上次的結果
            #         rows = repeated_sheet.max_row
            #         self.video_file.total_revised_interrupt_count = rows - 5
            #         start_frame = repeated_sheet['B']
            #         end_frame = repeated_sheet['D']
            #         label = repeated_sheet['F']
                    
            #         for i in range(self.video_file.total_revised_interrupt_count):
            #             interrupt_info = {}
            #             interrupt_info["start_frame"] = int(start_frame[i+1].value)
            #             interrupt_info["start_time"] = round(interrupt_info["start_frame"]/self.video_fps, 1)
            #             interrupt_info["end_frame"] = int(end_frame[i+1].value)
            #             interrupt_info["end_time"] = round(interrupt_info["end_frame"]/self.video_fps, 1)
            #             interrupt_info["label"] = label[i+1].value
            #             self.video_file.revised_interrupt_list.append(interrupt_info)          
        
            self.ui.button_start_judge.setDisabled(False)

    
    ### Start Judge
    def clicked_start_judge(self):

        # if self.first_file is True: ## 尚未第一次進入
        #     print("選video!!")
        #     mbox = QtWidgets.QMessageBox(self.ui.centralwidget) # 跳出警告訊息
        #     mbox.setIcon(QtWidgets.QMessageBox.Warning)
        #     mbox.setText("請先選擇要偵測的手術片段")
        #     mbox.setStandardButtons(QtWidgets.QMessageBox.Ok)
        #     mbox.exec()
        # else:
            for video in self.one_day_video:
                ##### Step1. start judge #####
                judge.start_judge(file=video)
                judge.revise_interrupt(file=video)


                ##### Step2. write result to file #####
                # video.write_result_to_file()


                ##### Step3. write result to excel #####
                video.write_result_to_excel(self.wb, self.dir_name)  


            ##### Step4. print result to UI #####
            # record interrupt times, wrong judge times, accuracy
            # self.wrongJudgeTimes = 0        # 目前沒這個資訊
            # self.accuracy = 0               # 有了total_interrupt_count跟wrongJudgeTimes就可算accuracy
            # self.ui.label_interrupt_times.setText("Interrupt Times:   " + str(self.video_file.total_revised_interrupt_count))

            self.movie = QtGui.QMovie("image/giphy.gif")
            self.ui.label_done.setMovie(self.movie)
            self.movie.start()

            self.ui.button_check_result.setDisabled(False)

            # song = AudioSegment.from_mp3("sound/done.mp3")
            # play(song)

    def show_result(self):
        self.ui.comboBox_choose_video.clear()
        for video_file in self.one_day_video:
            self.ui.comboBox_choose_video.addItem(video_file.filename)

    def change_list_widget(self):
        video_file_name = self.ui.comboBox_choose_video.currentText()

        ##### Step1. check現在是哪一個video #####
        index = self.ui.comboBox_choose_video.currentIndex()
        self.ui.list_widget_interrupt.clear() # 清空listwidget

        ##### Step2. show result in list widget #####
        self.choose_video_file = self.one_day_video[index]
        for i in range(self.choose_video_file.total_revised_interrupt_count):
            start_normal_time = compute.get_normal_time_info(time_in_seconds=self.choose_video_file.revised_interrupt_list[i]["start_time"])
            start_time_name = compute.get_excel_str(normal_time=start_normal_time)

            end_normal_time = compute.get_normal_time_info(time_in_seconds=self.choose_video_file.revised_interrupt_list[i]["end_time"])
            end_time_name = compute.get_excel_str(normal_time=end_normal_time)

            self.ui.list_widget_interrupt.addItem("Interrupt" + str(i+1) + ". " + start_time_name + " - " + end_time_name)
    
    def show_interrupt_clip(self):
        ##### Step1. 取出選取到的interrupt 
        item_index = self.ui.list_widget_interrupt.currentRow()

        start_choose_frame = self.choose_video_file.revised_interrupt_list[item_index]['start_frame']
        end_choose_frame = self.choose_video_file.revised_interrupt_list[item_index]['end_frame']

        ##### Step2. 播映interrupt開始的地方
        frame_diff = end_choose_frame - start_choose_frame
        count_frame = 0

        videoinfo = opencv_engine.get_video_info(self.choose_video_file.filepath)
        vc = videoinfo["vc"]
        fps = videoinfo["fps"]
        vc.set(cv.CAP_PROP_POS_FRAMES, start_choose_frame - fps*1) #往前1秒開始播放
        while True:
            count_frame += 1
            ret, frame = vc.read()             # 讀取影片的每一幀
            if not ret:
                print("Cannot receive frame")   # 如果讀取錯誤，印出訊息
                break
            cv.imshow('Interrupy fragment', frame)     # 如果讀取成功，顯示該幀的畫面
            if cv.waitKey(int(1000/fps)) == ord('q') or count_frame == int(frame_diff + fps*1): # 往後1秒結束(變成int才能偵測幀數)
                break
        
        vc.release()
        cv.destroyAllWindows()                 # 結束所有視窗
        

    def choose_remove_interrupt(self):
        self.remove_item_index = self.ui.list_widget_interrupt.currentRow()


    def remove_interrupt(self):
        # 刪除list widget的item
        remove_item = self.ui.list_widget_interrupt.takeItem(self.remove_item_index)
        self.ui.list_widget_interrupt.removeItemWidget(remove_item)

        # 刪除excel裡的row
        self.choose_video_file.delete_excel_row(self.wb, self.dir_name, self.remove_item_index)


    def performance(self):
        filePath, filterType = QtWidgets.QFileDialog.getOpenFileNames()  # 選取多個檔案
        wb = openpyxl.load_workbook('Result.xlsx')
        one_day_sheets = []
        filename = []
        each_interrupt_number = []
        one_day_interrupt_number = 0
        for i in range(len(filePath)):
            basename = os.path.basename(filePath[i])
            filename.append(os.path.splitext(basename)[0])
            one_day_sheets.append(wb[filename[i]])
            rows = one_day_sheets[i].max_row
            each_interrupt_number.append(rows - 5)
            one_day_interrupt_number += each_interrupt_number[i]
        video_total_time = 7383
        
        score = one_day_interrupt_number/(video_total_time/60/9)
        performance = '-'
        if 0 <= score < 1:
            performance = 'A+'
        elif 1 <= score < 3:
            performance = 'A'
        elif 3 <= score < 4:
            performance = 'A-'
        elif 4 <= score < 5:
            performance = 'B+'
        elif 5 <= score < 7:
            performance = 'B'
        elif 7 <= score < 8:
            performance = 'B-'
        elif 8 <= score < 9:
            performance = 'C+'
        elif 9 <= score < 11:
            performance = 'C'
        elif 11 <= score < 12:
            performance = 'C-'
        elif 12 <= score:
            performance = 'D'
        self.ui.label_performance.setText(str(performance))
            
        